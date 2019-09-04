"""script takes file in csv with header:
"Job","Clocked In","Clocked Out","Duration","Hourly Rate","Earnings","Comment",
"Tags","Breaks","Adjustments","TotalTimeAdjustment","TotalEarningsAdjustment"
and create new file with columns:
"Star","End","Total","Hours","Minutes"
"""

import sys
from numpy import array
from numpy import delete
import os
from datetime import date
from openpyxl import load_workbook


def load(fileName):
    if DEBUG: print(fileName)
    with open(fileName) as file:
        file.readline()
        file.readline()
        return file.readlines()


def parse(dateList):
    return [line.split('","') for line in dateList]


def shorten(dateList):
    return delete(array([array(line) for line in dateList]), [0, 4, 5, 6, 7, 8, 9, 10, 11], 1).tolist()


def reshape_rows(dateList):
    for line in dateList:
        try:
            line.append(line[2][:line[2].index(",", )])
        except ValueError:
            line.append(line[2])
        try:
            line.append(str((float("0" + "".join(line[2][line[2].index(",", ):]).replace(",", ".")) * 60).__round__()))
        except ValueError:
            line.append("00")
    return dateList


def reshape_cols(dateList):
    res = [[] for i in range(31)]
    for line in dateList:
        res[int(line[0][:2]) - 1] = line
    return res


def strip_date(dateList):
    for line in dateList:
        try:
            line[0] = line[0][-5:]
            line[1] = line[1][-5:]
        except IndexError:
            pass
    return dateList


def create_table(fileName):
    return strip_date(reshape_cols(reshape_rows(shorten(parse(load(fileName))))))


def write_xlsx(file, dateList):
    # loading xlsx and selecting right sheet
    wb = load_workbook(filename=file)
    ws = wb.active

    # writing table
    r, c = 1, 1
    for row in dateList:
        for item in row:
            if c == 3:
                if DEBUG: print("writing in cell: " + str(r + 9) + ", " + str(c + 1))
                ws.cell(r + 9, c + 1).value = float(item.replace(",", "."))
            else:
                if DEBUG: print("writing in cell: " + str(r + 9) + ", " + str(c + 1))
                ws.cell(r + 9, c + 1).value = str(item)
            c += 1
        c = 1
        r += 1

    # save updated *.xlsx to project root and save one to history
    wb.save(os.path.join(PROJECT_PATH, XLSX_FILE))
    wb.save(os.path.join(PROJECT_PATH, "hist", date.today().strftime("%d-%m-%Y") + "__" + XLSX_FILE))


def run(inputFile):
    write_xlsx(XLSX_FILE_PATH, create_table(inputFile))


def print_list(l):
    for line in l:
        print([str(item) for item in line])


if __name__ == '__main__':
    DEBUG = False

    PROJECT_PATH = os.path.abspath('.')
    XLSX_FILE = "Dochazkovy_list_BRIGADNICI.xlsx"
    XLSX_FILE_PATH = os.path.join(PROJECT_PATH, "bin", XLSX_FILE)
    SCRIPT_PATH = os.path.join(PROJECT_PATH, sys.argv[1])

    if DEBUG: print(PROJECT_PATH)
    if DEBUG:
        run(SCRIPT_PATH)
    else:
        try:
            run(sys.argv[1])
            print("DONE")
        except:
            print("some errors ocured, file was not created")
            os.system("pause")
